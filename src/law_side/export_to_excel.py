"""Export extracted/interim artifacts to Excel for human review."""

from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

from law_side.law_rulebase_models import (
    LegalDocument,
    LegalFrame,
    LegalUnit,
    NormativeSentence,
    PredicateLexiconEntry,
    RuleSeed,
)
from utils.logger import get_logger


_MANIFEST_COLUMNS: list[str] = [
    "doc_id",
    "doc_code",
    "doc_title",
    "doc_type",
    "issuing_body",
    "issue_date",
    "effective_date",
    "source_file_name",
    "source_format",
    "domain_scope",
    "domain_subscope",
    "document_role",
    "expected_rule_density",
    "parse_strategy",
    "is_consolidated_version",
    "amends_doc_codes",
    "has_appendix_forms",
    "legal_scope_note",
    "priority",
    "status",
    "notes",
]

_LEGAL_UNITS_COLUMNS: list[str] = [
    "unit_id",
    "doc_id",
    "doc_code",
    "chapter",
    "section",
    "article",
    "clause",
    "point",
    "unit_type",
    "unit_ref_full",
    "sentence_index",
    "subsentence_index",
    "list_item_marker",
    "heading",
    "text",
    "parent_context",
    "deontic_signal",
    "topic_tag",
    "normative_signal",
    "has_condition_marker",
    "has_deadline_marker",
    "has_document_marker",
    "has_authority_marker",
    "has_exception_marker",
    "has_threshold_marker",
    "has_cross_reference",
    "cross_reference_text",
    "actor_hint",
    "action_hint",
    "object_hint",
    "rule_density_estimate",
    "needs_split",
    "split_reason",
    "is_candidate_rule_sentence",
    "source_ref",
    "notes",
]

_CANDIDATE_NS_COLUMNS: list[str] = [
    "candidate_id",
    "unit_id",
    "doc_id",
    "doc_code",
    "unit_ref_full",
    "source_ref",
    "heading",
    "parent_context",
    "source_text",
    "sentence_text",
    "candidate_type",
    "candidate_subtype",
    "candidate_score",
    "trigger_patterns",
    "actor_text",
    "action_text",
    "object_text",
    "condition_text",
    "deadline_text",
    "authority_text",
    "document_text",
    "exception_text",
    "threshold_text",
    "legal_effect_text",
    "should_extract_rule",
    "extraction_priority",
    # Keep original fields for traceability/debug (but treated as meta).
    "sentence_type",
    "normative_pattern",
    "subject_span",
    "action_span",
    "modality_span",
    "condition_span",
    "time_span",
    "document_span",
    "authority_span",
    "candidate_rule_type",
    "confidence_manual",
    "ns_id",
    "notes",
]

_LEGAL_FRAMES_COLUMNS: list[str] = [
    "frame_id",
    "candidate_id",
    "source_unit_id",
    "doc_id",
    "doc_code",
    "unit_ref_full",
    "source_ref",
    "heading",
    "parent_context",
    "source_text",
    "frame_type",
    "chu_the",
    "vai_tro_chu_the",
    "hanh_vi",
    "doi_tuong_hanh_vi",
    "tinh_chat_phap_ly",
    "dieu_kien_ap_dung",
    "dieu_kien_dinh_luong",
    "nguong_so_luong",
    "nguong_ty_le",
    "khoang_gia_tri",
    "thanh_phan_ho_so",
    "co_quan_tiep_nhan",
    "co_quan_xu_ly",
    "ket_qua_thu_tuc",
    "thoi_han_so",
    "don_vi_thoi_han",
    "moc_tinh_thoi_han",
    "ngoai_le",
    "van_ban_dan_chieu",
    "ghi_chu_giai_thich",
    "muc_do_day_du",
    "can_tach_them",
    "ly_do_can_tach",
]

_PREDICATE_LEXICON_COLUMNS: list[str] = [
    "predicate_id",
    "surface_form",
    "bien_the_ngon_ngu",
    "hanh_vi_chuan",
    "hanh_vi_chuan_chi_tiet",
    "nhom_hanh_vi",
    "chu_the_mac_dinh",
    "doi_tuong_mac_dinh",
    "co_quan_mac_dinh",
    "can_thoi_han",
    "can_ho_so",
    "can_ngoai_le",
    "can_nguong_dinh_luong",
    "ghi_chu_ap_dung",
]

_RULEBASE_SEED_COLUMNS: list[str] = [
    # A. Nhóm nhận diện
    "rule_id",
    "rule_group_id",
    "frame_id",
    "candidate_id",
    "source_unit_id",
    "doc_id",
    "doc_code",
    # B. Nhóm truy vết nguồn
    "source_ref",
    "source_ref_full",
    "heading",
    "parent_context",
    "source_text",
    # C. Nhóm phân loại rule
    "rule_type",
    "tinh_chat_phap_ly",
    "canonical_predicate",
    "typed_predicate",
    "predicate_family",
    # D. Nhóm nội dung pháp lý cốt lõi
    "chu_the",
    "loai_chu_the",
    "vai_tro_chu_the",
    "dieu_kien_ap_dung",
    "bieu_thuc_dieu_kien",
    "hanh_vi_phap_ly",
    "doi_tuong_hanh_vi",
    "he_qua_phap_ly",
    # E. Nhóm định lượng / ngưỡng
    "ten_chi_so",
    "toan_tu_so_sanh",
    "gia_tri_nguong",
    "don_vi_nguong",
    "gia_tri_tu",
    "gia_tri_den",
    "kieu_khoang",
    # F. Nhóm thời hạn
    "thoi_han_so",
    "don_vi_thoi_han",
    "moc_tinh_thoi_han",
    "bieu_thuc_thoi_han",
    # G. Nhóm thủ tục / hồ sơ / cơ quan
    "thanh_phan_ho_so",
    "co_quan_tiep_nhan",
    "co_quan_xu_ly",
    "ket_qua_thu_tuc",
    "phuong_thuc_thuc_hien",
    # H. Nhóm ngoại lệ / phạm vi
    "pham_vi_ap_dung",
    "ngoai_le",
    "van_ban_dan_chieu",
    # I. Nhóm phục vụ answer / explanation
    "answer_template",
    "explanation_template",
    "grounded_summary",
    # J. Nhóm chất lượng
    "muc_do_day_du",
    "do_tin_cay_trich_xuat",
    "can_ra_soat",
    "ly_do_can_ra_soat",
    "extraction_pattern",
    "notes",
]


_EXCEL_ILLEGAL_CHARS_RE = None


def _map_meta_value(value: Any, mapping: dict[str, str]) -> Any:
    if value is None:
        return value
    s = str(value).strip()
    if not s:
        return value
    return mapping.get(s, value)


def _apply_grounded_meta_mappings(
    df_units: pd.DataFrame | None = None,
    df_ns: pd.DataFrame | None = None,
    df_frames: pd.DataFrame | None = None,
) -> None:
    def _to_co_khong(v: Any) -> Any:
        if isinstance(v, bool):
            return "co" if v else "khong"
        if v is None:
            return v
        s = str(v).strip().lower()
        if s in {"true", "1", "yes"}:
            return "co"
        if s in {"false", "0", "no"}:
            return "khong"
        return v

    unit_type_map = {
        "article": "dieu",
        "clause": "khoan",
        "point": "diem",
    }
    topic_tag_map = {
        "enterprise_registration": "dang_ky_doanh_nghiep",
        "enterprise_change": "thay_doi_dang_ky_doanh_nghiep",
        "branch_office": "chi_nhanh_van_phong_dai_dien",
        "branch_office_registration": "chi_nhanh_van_phong_dai_dien",
        "beneficial_owner": "chu_so_huu_huong_loi",
        "dossier": "ho_so",
        "document_requirement": "ho_so",
        "deadline": "thoi_han",
        "authority_action": "hanh_dong_co_quan",
        "enterprise_name": "ten_doanh_nghiep",
    }
    normative_signal_map = {
        "must": "phai",
        "shall": "phai",
        "must_not": "khong_duoc",
        "prohibition": "khong_duoc",
        "responsibility": "co_trach_nhiem",
        "has_responsibility": "co_trach_nhiem",
        "liability": "chiu_trach_nhiem",
        "obligation": "co_nghia_vu",
        "within_time_limit": "trong_thoi_han",
        "deadline": "trong_thoi_han",
        "dossier_includes": "ho_so_bao_gom",
        "dossier_contains": "ho_so_bao_gom",
        "notify": "thong_bao",
        "notification": "thong_bao",
        "register": "dang_ky",
        "registration": "dang_ky",
        "issue": "cap",
        "update": "cap_nhat",
        "revoke": "thu_hoi",
    }
    sentence_type_map = {
        "duty": "nghia_vu",
        "obligation": "nghia_vu",
        "prohibition": "cam",
        "permission": "quyen",
        "deadline": "thoi_han",
        "condition": "dieu_kien",
        "document_requirement": "ho_so",
        "document_rule": "ho_so",
        "authority_action": "hanh_dong_co_quan",
        "procedure": "thu_tuc",
        "procedure_rule": "thu_tuc",
        "other": "khac",
        "article": "dieu",
        "clause": "khoan",
        "point": "diem",
    }
    normative_pattern_map = {
        "subject_must_act": "chu_the_phai_hanh_dong",
        "subject_must_not_act": "chu_the_khong_duoc_hanh_dong",
        "subject_may_act": "chu_the_co_quyen_hanh_dong",
        "within_x_days": "trong_thoi_han_x_ngay",
        "dossier_includes": "ho_so_bao_gom",
        "authority_must_process": "co_quan_co_trach_nhiem",
        "notice_must_include": "thong_bao_phai_bao_gom",
        "obligation": "nghia_vu",
        "prohibition": "cam",
        "permission": "quyen",
        "deadline": "thoi_han",
        "dossier": "ho_so",
        "procedure": "thu_tuc",
    }
    candidate_rule_type_map = {
        "duty_rule": "quy_pham_nghia_vu",
        "duty": "quy_pham_nghia_vu",
        "registration_obligation": "quy_pham_nghia_vu",
        "prohibition_rule": "quy_pham_cam",
        "prohibition": "quy_pham_cam",
        "permission_rule": "quy_pham_quyen",
        "permission": "quy_pham_quyen",
        "deadline_rule": "quy_pham_thoi_han",
        "deadline": "quy_pham_thoi_han",
        "document_rule": "quy_pham_ho_so",
        "document_requirement": "quy_pham_ho_so",
        "authority_action": "hanh_dong_co_quan",
        "procedure_rule": "quy_pham_thu_tuc",
        "procedure": "quy_pham_thu_tuc",
        "condition": "quy_pham_dieu_kien",
        "condition_rule": "quy_pham_dieu_kien",
        "other": "khac",
    }
    confidence_map = {
        "high": "cao",
        "medium": "trung_binh",
        "low": "thap",
    }
    frame_type_map = {
        "duty_rule": "quy_pham_nghia_vu",
        "document_rule": "quy_pham_ho_so",
        "authority_action": "hanh_dong_co_quan",
        "procedure_rule": "quy_pham_thu_tuc",
        "condition_rule": "quy_pham_dieu_kien",
        "permission_rule": "quy_pham_quyen",
        "permission": "quy_pham_quyen",
        "prohibition_rule": "quy_pham_cam",
        "prohibition": "quy_pham_cam",
        "status_rule": "quy_pham_trang_thai",
        "quy phạm nghĩa vụ": "quy_pham_nghia_vu",
        "hành động của cơ quan": "hanh_dong_co_quan",
        "quy phạm thủ tục": "quy_pham_thu_tuc",
        "quy phạm hồ sơ": "quy_pham_ho_so",
        "quy phạm điều kiện": "quy_pham_dieu_kien",
    }
    modality_map = {
        "obligation": "nghia_vu",
        "permission": "quyen",
        "prohibition": "cam",
        "authority_action": "hanh_dong_co_quan",
        "procedure": "thu_tuc",
        "other": "khac",
        "nghĩa vụ": "nghia_vu",
        "quyền": "quyen",
        "cấm": "cam",
    }
    output_status_map = {
        "kept": "giu_lai",
        "keep": "giu_lai",
        "seed_extracted_first_pass": "giu_lai",
        "review": "can_ra_soat",
        "needs_review": "can_ra_soat",
        "low_confidence": "do_tin_cay_thap",
        "dropped": "loai",
        "drop": "loai",
    }
    subject_role_map = {
        "actor": "chu_the_thuc_hien",
        "regulated_subject": "chu_the_bi_dieu_chinh",
        "authority": "co_quan",
        "applicant": "chu_the_de_nghi",
        "enterprise_subject": "chu_the_doanh_nghiep",
    }

    if df_units is not None and not df_units.empty:
        if "unit_type" in df_units.columns:
            df_units["unit_type"] = df_units["unit_type"].map(lambda v: _map_meta_value(v, unit_type_map))
        if "topic_tag" in df_units.columns:
            df_units["topic_tag"] = df_units["topic_tag"].map(lambda v: _map_meta_value(v, topic_tag_map))
        if "normative_signal" in df_units.columns:
            df_units["normative_signal"] = df_units["normative_signal"].map(
                lambda v: _map_meta_value(v, normative_signal_map)
            )
        # Marker columns exported as co/khong for reviewer-friendly layer.
        for col in [
            "has_condition_marker",
            "has_deadline_marker",
            "has_document_marker",
            "has_authority_marker",
            "has_exception_marker",
            "has_threshold_marker",
            "has_cross_reference",
        ]:
            if col in df_units.columns:
                df_units[col] = df_units[col].map(_to_co_khong)

    if df_ns is not None and not df_ns.empty:
        if "sentence_type" in df_ns.columns:
            df_ns["sentence_type"] = df_ns["sentence_type"].map(lambda v: _map_meta_value(v, sentence_type_map))
        if "normative_pattern" in df_ns.columns:
            df_ns["normative_pattern"] = df_ns["normative_pattern"].map(
                lambda v: _map_meta_value(v, normative_pattern_map)
            )
        if "candidate_rule_type" in df_ns.columns:
            df_ns["candidate_rule_type"] = df_ns["candidate_rule_type"].map(
                lambda v: _map_meta_value(v, candidate_rule_type_map)
            )
        if "confidence_manual" in df_ns.columns:
            df_ns["confidence_manual"] = df_ns["confidence_manual"].map(
                lambda v: _map_meta_value(v, confidence_map)
            )

    if df_frames is not None and not df_frames.empty:
        if "frame_type" in df_frames.columns:
            df_frames["frame_type"] = df_frames["frame_type"].map(lambda v: _map_meta_value(v, frame_type_map))
        if "modality" in df_frames.columns:
            df_frames["modality"] = df_frames["modality"].map(lambda v: _map_meta_value(v, modality_map))
        if "output_status" in df_frames.columns:
            df_frames["output_status"] = df_frames["output_status"].map(
                lambda v: _map_meta_value(v, output_status_map)
            )
        if "subject_role" in df_frames.columns:
            df_frames["subject_role"] = df_frames["subject_role"].map(
                lambda v: _map_meta_value(v, subject_role_map)
            )


def _sanitize_for_excel(value: Any) -> Any:
    """Remove control characters that openpyxl cannot write to worksheets."""
    if value is None:
        return None
    if not isinstance(value, str):
        return value

    # openpyxl rejects ASCII control chars in range [0x00-0x1F] except \t,\n,\r.
    # Keep this conservative to preserve readability.
    import re

    # Lazily compile regex once.
    global _EXCEL_ILLEGAL_CHARS_RE
    if _EXCEL_ILLEGAL_CHARS_RE is None:
        _EXCEL_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

    return _EXCEL_ILLEGAL_CHARS_RE.sub("", value)


def _autosize_columns(xlsx_path: Path, columns: Iterable[str]) -> None:
    """Best-effort autosize based on string lengths."""
    try:
        from openpyxl import load_workbook
    except Exception:  # pragma: no cover
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Column names row assumed at header=1.
    header = {cell.value: cell.column for cell in ws[1] if cell.value is not None}

    for col_name in columns:
        if col_name not in header:
            continue
        col_idx = header[col_name]
        max_len = 8
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, min(60, len(str(val))))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len

    wb.save(xlsx_path)


def export_all_to_excel(
    *,
    documents: list[LegalDocument],
    legal_units: list[LegalUnit],
    normative_sentences: list[NormativeSentence],
    legal_frames: list[LegalFrame],
    predicate_lexicon: list[PredicateLexiconEntry],
    rule_seeds: list[RuleSeed],
    out_document_manifest: Path,
    out_legal_units: Path,
    out_candidate_sentences: Path,
    out_legal_frames: Path,
    out_predicate_lexicon: Path,
    out_rulebase_seed: Path,
    autosize: bool = False,
) -> None:
    """Export the 6 required Excel artifacts."""
    log = get_logger("export_to_excel")

    # Ensure directories.
    for p in [
        out_document_manifest,
        out_legal_units,
        out_candidate_sentences,
        out_legal_frames,
        out_predicate_lexicon,
        out_rulebase_seed,
    ]:
        p.parent.mkdir(parents=True, exist_ok=True)

    df_manifest = pd.DataFrame(
        [
            {k: _sanitize_for_excel(getattr(d, k)) for k in _MANIFEST_COLUMNS}
            for d in documents
        ],
        columns=_MANIFEST_COLUMNS,
    )
    df_manifest.to_excel(out_document_manifest, index=False)
    if autosize:
        _autosize_columns(out_document_manifest, _MANIFEST_COLUMNS)

    df_units = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(u, k)) for k in _LEGAL_UNITS_COLUMNS} for u in legal_units],
        columns=_LEGAL_UNITS_COLUMNS,
    )
    df_units.to_excel(out_legal_units, index=False)
    if autosize:
        _autosize_columns(out_legal_units, _LEGAL_UNITS_COLUMNS)

    df_ns = pd.DataFrame(
        [
            {k: _sanitize_for_excel(getattr(ns, k)) for k in _CANDIDATE_NS_COLUMNS}
            for ns in normative_sentences
        ],
        columns=_CANDIDATE_NS_COLUMNS,
    )
    df_ns.to_excel(out_candidate_sentences, index=False)
    if autosize:
        _autosize_columns(out_candidate_sentences, _CANDIDATE_NS_COLUMNS)

    df_frames = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(f, k)) for k in _LEGAL_FRAMES_COLUMNS} for f in legal_frames],
        columns=_LEGAL_FRAMES_COLUMNS,
    )
    df_frames.to_excel(out_legal_frames, index=False)
    if autosize:
        _autosize_columns(out_legal_frames, _LEGAL_FRAMES_COLUMNS)

    df_lex = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(e, k)) for k in _PREDICATE_LEXICON_COLUMNS} for e in predicate_lexicon],
        columns=_PREDICATE_LEXICON_COLUMNS,
    )
    df_lex.to_excel(out_predicate_lexicon, index=False)
    if autosize:
        _autosize_columns(out_predicate_lexicon, _PREDICATE_LEXICON_COLUMNS)

    df_rules = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(r, k)) for k in _RULEBASE_SEED_COLUMNS} for r in rule_seeds],
        columns=_RULEBASE_SEED_COLUMNS,
    )
    df_rules.to_excel(out_rulebase_seed, index=False)
    if autosize:
        _autosize_columns(out_rulebase_seed, _RULEBASE_SEED_COLUMNS)

    log.info("Exported 6 Excel files.")


def export_document_manifest(
    *,
    documents: list[LegalDocument],
    out_document_manifest: Path,
    autosize: bool = False,
) -> None:
    """Export only `document_manifest.xlsx`."""
    out_document_manifest.parent.mkdir(parents=True, exist_ok=True)
    df_manifest = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(d, k)) for k in _MANIFEST_COLUMNS} for d in documents],
        columns=_MANIFEST_COLUMNS,
    )
    df_manifest.to_excel(out_document_manifest, index=False)
    if autosize:
        _autosize_columns(out_document_manifest, _MANIFEST_COLUMNS)


def export_legal_units_review(
    *,
    legal_units: list[LegalUnit],
    out_legal_units: Path,
    autosize: bool = False,
) -> None:
    """Export only `legal_units_review.xlsx`."""
    out_legal_units.parent.mkdir(parents=True, exist_ok=True)
    df_units = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(u, k)) for k in _LEGAL_UNITS_COLUMNS} for u in legal_units],
        columns=_LEGAL_UNITS_COLUMNS,
    )
    _apply_grounded_meta_mappings(df_units=df_units)
    df_units.to_excel(out_legal_units, index=False)
    if autosize:
        _autosize_columns(out_legal_units, _LEGAL_UNITS_COLUMNS)


def export_candidate_normative_sentences(
    *,
    normative_sentences: list[NormativeSentence],
    out_candidate_sentences: Path,
    autosize: bool = False,
) -> None:
    """Export only `candidate_normative_sentences.xlsx`."""
    out_candidate_sentences.parent.mkdir(parents=True, exist_ok=True)
    df_ns = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(ns, k)) for k in _CANDIDATE_NS_COLUMNS} for ns in normative_sentences],
        columns=_CANDIDATE_NS_COLUMNS,
    )
    _apply_grounded_meta_mappings(df_ns=df_ns)
    df_ns.to_excel(out_candidate_sentences, index=False)
    if autosize:
        _autosize_columns(out_candidate_sentences, _CANDIDATE_NS_COLUMNS)


def export_legal_frames_review(
    *,
    legal_frames: list[LegalFrame],
    out_legal_frames: Path,
    autosize: bool = False,
) -> None:
    """Export only `legal_frames_review.xlsx`."""
    out_legal_frames.parent.mkdir(parents=True, exist_ok=True)
    df_frames = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(f, k)) for k in _LEGAL_FRAMES_COLUMNS} for f in legal_frames],
        columns=_LEGAL_FRAMES_COLUMNS,
    )
    _apply_grounded_meta_mappings(df_frames=df_frames)
    df_frames.to_excel(out_legal_frames, index=False)
    if autosize:
        _autosize_columns(out_legal_frames, _LEGAL_FRAMES_COLUMNS)


def export_predicate_lexicon(
    *,
    predicate_lexicon: list[PredicateLexiconEntry],
    out_predicate_lexicon: Path,
    autosize: bool = False,
) -> None:
    """Export only `predicate_lexicon.xlsx`."""
    out_predicate_lexicon.parent.mkdir(parents=True, exist_ok=True)
    df_lex = pd.DataFrame(
        [{k: _sanitize_for_excel(getattr(e, k)) for k in _PREDICATE_LEXICON_COLUMNS} for e in predicate_lexicon],
        columns=_PREDICATE_LEXICON_COLUMNS,
    )
    df_lex.to_excel(out_predicate_lexicon, index=False)
    if autosize:
        _autosize_columns(out_predicate_lexicon, _PREDICATE_LEXICON_COLUMNS)


__all__ = [
    "export_all_to_excel",
    "export_document_manifest",
    "export_legal_units_review",
    "export_candidate_normative_sentences",
    "export_legal_frames_review",
    "export_predicate_lexicon",
]

